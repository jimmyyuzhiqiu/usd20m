import os
import subprocess
import sys
import threading
from typing import Optional, Dict, List, Set, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# =========================
# 默认参数
# =========================
THRESHOLD_USD = 20_000_000
OUTPUT_SHEET_NAME = "USD_over_20M"
RATE_SHEET_NAME = "rate"

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ICON_PATH = os.path.join(BASE_DIR, "app.ico")
LOGO_SRC_PATH = os.path.join(BASE_DIR, "ing-logo.png")


# =========================
# 工具函数
# =========================
def parse_amount_input(text: str) -> Optional[float]:
    if text is None:
        return None
    raw = str(text).strip()
    if not raw:
        return None

    s = raw.upper().replace(",", "").replace(" ", "")
    multiplier = 1.0
    if s.endswith("B"):
        multiplier = 1_000_000_000.0
        s = s[:-1]
    elif s.endswith("M"):
        multiplier = 1_000_000.0
        s = s[:-1]
    elif s.endswith("K"):
        multiplier = 1_000.0
        s = s[:-1]

    try:
        value = float(s) * multiplier
    except ValueError:
        return None

    if value <= 0:
        return None
    return value


def parse_manual_rates(text: str) -> dict:
    """
    支持：
      JPY=156.9
      JPY:156.9
      JPY 156.9
    """
    if text is None:
        return {}
    lines = [line.strip() for line in str(text).splitlines()]
    rates = {}
    for line in lines:
        if not line:
            continue
        if "#" in line:
            line = line.split("#", 1)[0].strip()
            if not line:
                continue

        parts = None
        for sep in ("=", ":"):
            if sep in line:
                parts = [p.strip() for p in line.split(sep, 1)]
                break
        if parts is None:
            tokens = line.split()
            if len(tokens) >= 2:
                parts = [tokens[0], tokens[1]]

        if not parts or len(parts) < 2:
            continue

        ccy = parts[0].strip().upper()
        if len(ccy) != 3:
            continue

        val = parts[1].strip().replace(",", "").replace(" ", "")
        try:
            rate_val = float(val)
        except ValueError:
            continue

        if rate_val > 0:
            rates[ccy] = rate_val

    return rates


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]

    target_cols = {
        "Transfer Amount": ["Transfer Amount", "TransferAmount", "Transfer_Amount"],
        "SettleCurrency": ["SettleCurrency", "Settle Currency", "Settle_Currency"],
        "Product Type": ["Product Type", "ProductType", "Product_Type"],
        "Trade Id": ["Trade Id", "TradeId", "Trade_ID", "TradeID"],
        "Transfer Type": ["Transfer Type", "TransferType", "Transfer_Type"],
    }

    col_map_insensitive = {c.lower().replace(" ", "").replace("_", ""): c for c in df.columns}
    rename_dict = {}

    for std, candidates in target_cols.items():
        if std in df.columns:
            continue
        found = None
        for cand in candidates:
            key = cand.lower().replace(" ", "").replace("_", "")
            if key in col_map_insensitive:
                found = col_map_insensitive[key]
                break
        if found:
            rename_dict[found] = std

    if rename_dict:
        df = df.rename(columns=rename_dict)

    return df


def parse_amount_to_float(series: pd.Series) -> pd.Series:
    s = series.astype(str).str.replace(",", "", regex=False).str.replace(" ", "", regex=False)
    return pd.to_numeric(s, errors="coerce")


def apply_tradeid_rule_keep_all_none(df_filtered: pd.DataFrame) -> pd.DataFrame:
    df = df_filtered.copy()

    required = ["Product Type", "Trade Id", "Transfer Type"]
    if any(c not in df.columns for c in required):
        return df

    df["__row_order"] = range(len(df))

    mask = df["Product Type"].astype(str).str.strip().isin(["Cash", "StructuredFlows"])
    df_need = df[mask].copy()
    df_keep = df[~mask].copy()

    if df_need.empty:
        return df_keep.drop(columns=["__row_order"], errors="ignore")

    df_need["Trade Id"] = df_need["Trade Id"].astype(str)

    kept_parts = []
    for _, g in df_need.groupby("Trade Id", dropna=False, sort=False):
        g_sorted = g.sort_values("__row_order")
        g_none = g_sorted[g_sorted["Transfer Type"].astype(str).str.strip().str.upper() == "NONE"]
        if not g_none.empty:
            kept_parts.append(g_none)
        else:
            kept_parts.append(g_sorted.iloc[[0]])

    picked = pd.concat(kept_parts, ignore_index=False)

    out = pd.concat([df_keep, picked], ignore_index=True)
    out = out.sort_values("__row_order").drop(columns=["__row_order"], errors="ignore").reset_index(drop=True)
    return out


def autofit_column_width(ws, min_width=8, max_width=60, extra=2):
    col_max = {}
    for row in ws.iter_rows(values_only=True):
        for j, val in enumerate(row, start=1):
            length = 0 if val is None else len(str(val))
            col_max[j] = max(col_max.get(j, 0), length)

    for j, max_len in col_max.items():
        width = max(min_width, min(max_width, max_len + extra))
        ws.column_dimensions[get_column_letter(j)].width = width


def write_df_to_existing_workbook(file_path: str, sheet_name: str, df: pd.DataFrame):
    wb = load_workbook(file_path)

    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    ws = wb.create_sheet(title=sheet_name)

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    ws.freeze_panes = "A2"
    autofit_column_width(ws)

    wb.save(file_path)


# =========================
# rate 读取（重点：不依赖 LIMIT AMOUNT 公式结果）
# =========================
def _to_upper_str(x) -> str:
    return str(x).strip().upper()


def _to_num(x) -> Optional[float]:
    v = pd.to_numeric(str(x).replace(",", "").replace(" ", ""), errors="coerce")
    if pd.isna(v):
        return None
    vv = float(v)
    if vv <= 0:
        return None
    return vv


def _find_rate_sheet(file_path: str) -> str:
    xls = pd.ExcelFile(file_path, engine="openpyxl")
    for name in xls.sheet_names:
        if str(name).strip().lower() == RATE_SHEET_NAME:
            return name

    # fallback：扫描任意 sheet 前 40 行是否包含 CCY
    for name in xls.sheet_names:
        try:
            head = pd.read_excel(file_path, sheet_name=name, header=None, engine="openpyxl", nrows=40)
        except Exception:
            continue
        for _, row in head.iterrows():
            cells = [_to_upper_str(x) for x in row.tolist()]
            if "CCY" in cells:
                return name

    raise ValueError("未找到 rate sheet（无法识别含 CCY 表头的表）。")


def _find_threshold_usd(raw: pd.DataFrame, default_threshold: float) -> float:
    """
    在整张 raw 里找包含 USD 的行，从该行取“明显是阈值”的最大数字（>1000），避免取到 USD=1 那行。
    """
    best = None
    for _, row in raw.iterrows():
        cells = [_to_upper_str(x) for x in row.tolist()]
        if "USD" not in cells:
            continue
        nums = []
        for x in row.tolist():
            v = _to_num(x)
            if v is not None and v > 1000:
                nums.append(v)
        if nums:
            cand = max(nums)
            best = cand if best is None else max(best, cand)

    if best is not None:
        return float(best)
    return float(default_threshold)


def _detect_header_cols(raw: pd.DataFrame) -> Tuple[int, int, Optional[int], Optional[int]]:
    """
    返回：header_row_idx, ccy_col_idx, exchange_col_idx, limit_col_idx
    允许 “LIMIT” 与 “AMOUNT” 分开出现在相邻单元格的情况。
    """
    for i, row in raw.iterrows():
        cells = [_to_upper_str(x) for x in row.tolist()]
        if "CCY" not in cells:
            continue

        ccy_col = cells.index("CCY")

        exch_col = None
        limit_col = None

        for j, c in enumerate(cells):
            if exch_col is None and ("EXCHANGE" in c and "RATE" in c):
                exch_col = j

        # LIMIT AMOUNT 可能是一个单元格，也可能拆成 LIMIT / AMOUNT
        for j, c in enumerate(cells):
            if "LIMIT" in c and "AMOUNT" in c:
                limit_col = j
                break
        if limit_col is None:
            for j, c in enumerate(cells):
                if c == "LIMIT" and j + 1 < len(cells) and cells[j + 1] == "AMOUNT":
                    limit_col = j
                    break
        if limit_col is None:
            for j, c in enumerate(cells):
                if "LIMIT" in c:
                    limit_col = j
                    break

        return int(i), int(ccy_col), exch_col, limit_col

    raise ValueError("rate 表格式无法识别：未找到 CCY 表头行。")


def load_fx_rates_from_rate_sheet(file_path: str, default_threshold_usd: float) -> tuple[dict, float, str, dict]:
    """
    读取：
      - threshold_usd（自动识别）
      - 汇率（外币/美元）
    规则：
      1) 优先用 EXCHANGE RATE VS CNY + CNY 行推导：
         EXCHANGE: 1 CCY = x CNY
         CNY 行的 EXCHANGE 视为：1 USD = x CNY
         => CCY per USD = (CNY per USD) / (CNY per CCY)
      2) 若无法推导，再用 LIMIT AMOUNT / threshold_usd 兜底（注意：LIMIT 若是公式且未落盘，可能为空）
    返回：rates, threshold_usd, rate_sheet_name, details
    """
    rate_sheet = _find_rate_sheet(file_path)
    raw = pd.read_excel(file_path, sheet_name=rate_sheet, header=None, engine="openpyxl")

    threshold_usd = _find_threshold_usd(raw, default_threshold_usd)

    header_row, ccy_col, exch_col, limit_col = _detect_header_cols(raw)

    exch_map: Dict[str, float] = {}
    limit_map: Dict[str, float] = {}

    # 读取 exchange/limit（允许下面新增币种，循环到底）
    for _, r in raw.iloc[header_row + 1 :].iterrows():
        ccy = _to_upper_str(r.iloc[ccy_col])
        if not ccy or ccy in ("CCY", "NAN") or len(ccy) != 3:
            continue

        if exch_col is not None:
            v = _to_num(r.iloc[exch_col])
            if v is not None:
                exch_map[ccy] = float(v)

        if limit_col is not None:
            v2 = _to_num(r.iloc[limit_col])
            if v2 is not None:
                limit_map[ccy] = float(v2)

    rates: Dict[str, float] = {}

    # 1) 优先用 exchange 推导（不依赖 LIMIT 公式结果）
    cny_per_usd = exch_map.get("CNY") if exch_map else None
    if exch_map and cny_per_usd and cny_per_usd > 0:
        for ccy, cny_per_ccy in exch_map.items():
            if ccy == "USD":
                rates["USD"] = 1.0
                continue
            if cny_per_ccy <= 0:
                continue
            rates[ccy] = float(cny_per_usd) / float(cny_per_ccy)
        rates.setdefault("USD", 1.0)
        source = "exchange_vs_cny"
    else:
        # 2) fallback：limit/threshold
        if not limit_map:
            raise ValueError("rate 表无法读取有效汇率：EXCHANGE 推导失败，且 LIMIT AMOUNT 无有效数值。")
        for ccy, amt in limit_map.items():
            rates[ccy] = float(amt) / float(threshold_usd)
        rates.setdefault("USD", 1.0)
        source = "limit_amount"

    details = {
        "source": source,
        "exchange_map": exch_map,
        "limit_map": limit_map,
        "threshold_usd_detected": threshold_usd,
        "rate_sheet": rate_sheet,
    }
    return rates, float(threshold_usd), rate_sheet, details


def prescan_missing_currencies(
    file_path: str, rate_sheet_name: str, fx_rate_local_per_usd: dict
) -> list[str]:
    xls = pd.ExcelFile(file_path, engine="openpyxl")
    data_sheets = [s for s in xls.sheet_names if s not in (rate_sheet_name, OUTPUT_SHEET_NAME)]
    sheet_names = data_sheets[:3]

    missing: Set[str] = set()
    for sh in sheet_names:
        try:
            df = pd.read_excel(file_path, sheet_name=sh, engine="openpyxl")
        except Exception:
            continue

        df = normalize_columns(df)
        if "SettleCurrency" not in df.columns:
            continue

        ccy_upper = df["SettleCurrency"].astype(str).str.upper().str.strip()
        for c in ccy_upper.unique().tolist():
            if not c or c == "NAN" or len(str(c).strip()) != 3:
                continue
            c = str(c).strip().upper()
            if c not in fx_rate_local_per_usd:
                missing.add(c)

    return sorted(missing)


# =========================
# 主处理逻辑
# =========================
def process_workbook(
    file_path: str,
    progress_cb=None,
    status_cb=None,
    rates_cb=None,
    missing_cb=None,
    fx_rate_override: Optional[dict] = None,
    threshold_override: Optional[float] = None,
    rate_sheet_override: Optional[str] = None,
    sheet_source_override: Optional[str] = None,
    suppress_missing_ccy: Optional[Set[str]] = None,
) -> int:
    def report_status(msg: str):
        if status_cb:
            status_cb(msg)

    def report_progress(step: int, total: int):
        if progress_cb:
            report_progress(step, total)

    report_status("准备开始...")

    if fx_rate_override is None:
        raise ValueError("内部错误：缺少 fx_rate_override。")

    fx_rate_local_per_usd = dict(fx_rate_override)
    threshold_usd = float(threshold_override or THRESHOLD_USD)
    rate_sheet = rate_sheet_override or RATE_SHEET_NAME
    sheet_source = sheet_source_override or "sheet"

    fx_rate_local_per_usd.setdefault("USD", 1.0)

    if rates_cb:
        rates_cb(
            {
                "rates": fx_rate_local_per_usd,
                "threshold_usd": threshold_usd,
                "sheet_name": rate_sheet,
                "source": sheet_source,
            }
        )

    xls = pd.ExcelFile(file_path, engine="openpyxl")
    data_sheets = [s for s in xls.sheet_names if s not in (rate_sheet, OUTPUT_SHEET_NAME)]
    sheet_names = data_sheets[:3]

    if not sheet_names:
        write_df_to_existing_workbook(file_path, OUTPUT_SHEET_NAME, pd.DataFrame())
        report_status("未找到可处理的数据表，已生成空结果。")
        if progress_cb:
            progress_cb(100)
        return 0

    total_steps = len(sheet_names) + 1
    if progress_cb:
        progress_cb(0)

    results = []
    missing_ccy_set: Set[str] = set()

    for idx, sh in enumerate(sheet_names, start=1):
        report_status(f"处理中：{sh}")
        df = pd.read_excel(file_path, sheet_name=sh, engine="openpyxl")
        df = normalize_columns(df)

        must_cols = ["Transfer Amount", "SettleCurrency", "Product Type", "Trade Id", "Transfer Type"]
        miss = [c for c in must_cols if c not in df.columns]
        if miss:
            if progress_cb:
                progress_cb(int(idx / total_steps * 100))
            continue

        amt_num = parse_amount_to_float(df["Transfer Amount"])
        abs_amt = amt_num.abs()

        ccy_upper = df["SettleCurrency"].astype(str).str.upper().str.strip()
        rate = ccy_upper.map(fx_rate_local_per_usd)

        missing_ccy = sorted(set(ccy_upper[rate.isna()].tolist()))
        missing_ccy = [c for c in missing_ccy if c and c != "NAN"]
        if missing_ccy:
            missing_ccy_set.update(missing_ccy)

        usd_amt = abs_amt / rate

        df_big = df[usd_amt > threshold_usd].copy()
        if df_big.empty:
            if progress_cb:
                progress_cb(int(idx / total_steps * 100))
            continue

        df_big = apply_tradeid_rule_keep_all_none(df_big)
        df_big.loc[:, "Transfer Amount"] = abs_amt.loc[df_big.index]
        results.append(df_big)

        if progress_cb:
            progress_cb(int(idx / total_steps * 100))

    report_status("写入结果...")

    if missing_ccy_set and missing_cb:
        sup = suppress_missing_ccy or set()
        to_warn = sorted(set(missing_ccy_set) - set(sup))
        if to_warn:
            missing_cb(to_warn)

    if results:
        out_df = pd.concat(results, ignore_index=True)
        write_df_to_existing_workbook(file_path, OUTPUT_SHEET_NAME, out_df)
        if progress_cb:
            progress_cb(100)
        return len(out_df)

    write_df_to_existing_workbook(file_path, OUTPUT_SHEET_NAME, pd.DataFrame())
    if progress_cb:
        progress_cb(100)
    return 0


def open_file_path(file_path: str) -> None:
    if not file_path or not os.path.exists(file_path):
        return
    try:
        if sys.platform.startswith("win"):
            os.startfile(file_path)  # type: ignore[attr-defined]
        elif sys.platform == "darwin":
            subprocess.run(["open", file_path], check=False)
        else:
            subprocess.run(["xdg-open", file_path], check=False)
    except Exception:
        pass


# =========================
# UI 相关
# =========================
def load_qt_modules():
    try:
        from PySide6 import QtCore, QtGui, QtWidgets

        return "PySide6", QtCore, QtGui, QtWidgets, QtCore.Signal, QtCore.Slot
    except Exception:
        pass

    try:
        from PyQt6 import QtCore, QtGui, QtWidgets

        return "PyQt6", QtCore, QtGui, QtWidgets, QtCore.pyqtSignal, QtCore.pyqtSlot
    except Exception:
        pass

    try:
        from PyQt5 import QtCore, QtGui, QtWidgets

        return "PyQt5", QtCore, QtGui, QtWidgets, QtCore.pyqtSignal, QtCore.pyqtSlot
    except Exception:
        return None


def launch_ui() -> bool:
    qt_modules = load_qt_modules()
    if not qt_modules:
        print("UI 依赖 Qt (PySide6 / PyQt6 / PyQt5)，当前环境未安装。")
        print("请先执行: python -m pip install PySide6")
        return False

    _qt_name, QtCore, QtGui, QtWidgets, Signal, Slot = qt_modules
    Qt = QtCore.Qt

    def get_qt_plugins_path() -> Optional[str]:
        qlib = getattr(QtCore, "QLibraryInfo", None)
        if qlib is None:
            return None
        lib_path = getattr(qlib, "LibraryPath", None)
        if lib_path and hasattr(qlib, "path"):
            return qlib.path(lib_path.PluginsPath)
        if hasattr(qlib, "location") and hasattr(qlib, "PluginsPath"):
            return qlib.location(qlib.PluginsPath)
        return None

    plugins_path = get_qt_plugins_path()
    if plugins_path:
        platform_path = os.path.join(plugins_path, "platforms")
        os.environ.setdefault("QT_PLUGIN_PATH", plugins_path)
        os.environ.setdefault("QT_QPA_PLATFORM_PLUGIN_PATH", platform_path)
        try:
            QtCore.QCoreApplication.addLibraryPath(plugins_path)
        except Exception:
            pass

    align_flag = getattr(Qt, "AlignmentFlag", None)
    if align_flag:
        align_left = align_flag.AlignLeft
        align_right = align_flag.AlignRight
        align_center = align_flag.AlignCenter
        align_vcenter = align_flag.AlignVCenter
    else:
        align_left = Qt.AlignLeft
        align_right = Qt.AlignRight
        align_center = Qt.AlignCenter
        align_vcenter = Qt.AlignVCenter

    no_pen = getattr(Qt, "NoPen", None)
    if no_pen is None:
        no_pen = Qt.PenStyle.NoPen

    wa_styled = getattr(Qt, "WA_StyledBackground", None)
    if wa_styled is None:
        wa_styled = Qt.WidgetAttribute.WA_StyledBackground

    smooth_transform = getattr(Qt, "SmoothTransformation", None)
    if smooth_transform is None:
        smooth_transform = Qt.TransformationMode.SmoothTransformation

    aspect_keep = getattr(Qt, "KeepAspectRatio", None)
    if aspect_keep is None:
        aspect_keep = Qt.AspectRatioMode.KeepAspectRatio

    render_hint = getattr(QtGui.QPainter, "Antialiasing", None)
    if render_hint is None:
        render_hint = QtGui.QPainter.RenderHint.Antialiasing

    def font_bold_weight():
        if hasattr(QtGui.QFont, "Weight"):
            return QtGui.QFont.Weight.Bold
        return QtGui.QFont.Bold

    def create_rounded_pixmap(path: str, size: int, radius: int):
        if not os.path.exists(path):
            return None

        pixmap = QtGui.QPixmap(path)
        if pixmap.isNull():
            return None

        scaled = pixmap.scaled(size, size, aspect_keep, smooth_transform)
        target = QtGui.QPixmap(size, size)
        target.fill(QtGui.QColor(0, 0, 0, 0))

        painter = QtGui.QPainter(target)
        painter.setRenderHint(render_hint, True)

        clip_path = QtGui.QPainterPath()
        clip_path.addRoundedRect(QtCore.QRectF(0, 0, size, size), radius, radius)
        painter.setClipPath(clip_path)

        x_offset = int((size - scaled.width()) / 2)
        y_offset = int((size - scaled.height()) / 2)
        painter.drawPixmap(x_offset, y_offset, scaled)
        painter.end()

        return target

    def is_pyqt6() -> bool:
        return _qt_name == "PyQt6"

    def non_editable_item(item):
        try:
            if is_pyqt6():
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            else:
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)
        except Exception:
            pass
        return item

    DIALOG_STYLE = """
        QDialog {
            background-color: #0F212B;
            border: 1px solid #173443;
            border-radius: 14px;
        }
        QLabel { color: #E6F1F5; }
        QLabel#hint { color: #9FB2BC; }
        QLineEdit {
            background: #0A1720;
            border: 1px solid #23414D;
            border-radius: 10px;
            padding: 8px 10px;
            color: #E6F1F5;
        }
        QLineEdit:focus { border-color: #44E6C2; }
        QTableWidget {
            background: #0A1720;
            border: 1px solid #23414D;
            border-radius: 10px;
            gridline-color: #23414D;
            color: #E6F1F5;
        }
        QHeaderView::section {
            background-color: #132833;
            color: #E6F1F5;
            padding: 6px;
            border: 1px solid #23414D;
        }
        QPushButton#primary {
            background-color: #44E6C2;
            color: #0A1B22;
            border-radius: 10px;
            padding: 8px 16px;
        }
        QPushButton#primary:hover { background-color: #3BD9B7; }
        QPushButton#secondary {
            background-color: #132833;
            color: #E6F1F5;
            border: 1px solid #23414D;
            border-radius: 10px;
            padding: 8px 16px;
        }
        QPushButton#secondary:hover { border-color: #44E6C2; }
    """

    class BackgroundWidget(QtWidgets.QWidget):
        def paintEvent(self, event):
            painter = QtGui.QPainter(self)
            gradient = QtGui.QLinearGradient(0, 0, 0, self.height())
            gradient.setColorAt(0, QtGui.QColor("#0B1F2A"))
            gradient.setColorAt(1, QtGui.QColor("#0E242E"))
            painter.fillRect(self.rect(), gradient)

            painter.setRenderHint(render_hint, True)
            painter.setPen(no_pen)
            painter.setBrush(QtGui.QColor("#123645"))
            painter.drawEllipse(-120, -80, 380, 380)

            painter.setBrush(QtGui.QColor("#123240"))
            painter.drawEllipse(self.width() - 260, self.height() - 200, 480, 480)

            pen = QtGui.QPen(QtGui.QColor("#1B3946"), 2)
            painter.setPen(pen)
            line_y = int(self.height() * 0.32)
            painter.drawLine(0, line_y, self.width(), line_y)

    class ThresholdConfirmDialog(QtWidgets.QDialog):
        """
        阈值：先显示自动识别值，再允许用户确认/修改。
        """
        def __init__(self, detected: float, current: float, rate_sheet: str, parent=None):
            super().__init__(parent)
            self.setWindowTitle("确认 USD 阈值")
            self.setMinimumSize(520, 260)
            self.setStyleSheet(DIALOG_STYLE)
            self._value = None

            layout = QtWidgets.QVBoxLayout(self)
            layout.setContentsMargins(16, 16, 16, 16)
            layout.setSpacing(10)

            title = QtWidgets.QLabel("USD 阈值确认")
            f = QtGui.QFont()
            f.setPointSize(13)
            f.setWeight(font_bold_weight())
            title.setFont(f)
            layout.addWidget(title)

            hint = QtWidgets.QLabel(
                f"已从 rate 表（{rate_sheet}）自动识别到阈值：{detected:,.2f}。\n"
                "你可以直接确认，也可以在下方修改后再继续导出。"
            )
            hint.setObjectName("hint")
            hint.setWordWrap(True)
            layout.addWidget(hint)

            self.edit = QtWidgets.QLineEdit()
            self.edit.setPlaceholderText("例如：20000000 或 20M")
            self.edit.setText(f"{current:,.2f}")
            layout.addWidget(self.edit)

            btn_row = QtWidgets.QHBoxLayout()
            btn_row.addStretch(1)

            self.cancel_btn = QtWidgets.QPushButton("取消")
            self.cancel_btn.setObjectName("secondary")
            self.ok_btn = QtWidgets.QPushButton("确认并继续")
            self.ok_btn.setObjectName("primary")

            btn_row.addWidget(self.cancel_btn)
            btn_row.addWidget(self.ok_btn)
            layout.addLayout(btn_row)

            self.cancel_btn.clicked.connect(self._on_cancel)
            self.ok_btn.clicked.connect(self._on_ok)

        def _on_cancel(self):
            self._value = None
            self.reject()

        def _on_ok(self):
            txt = self.edit.text().strip()
            v = parse_amount_input(txt)
            if v is None:
                # 允许用户直接输入带逗号的小数
                v = _to_num(txt)
            if v is None or v <= 0:
                QtWidgets.QMessageBox.warning(self, "提示", "阈值格式不正确，请输入正数（如 20M / 20000000）。")
                return
            self._value = float(v)
            self.accept()

        def value(self) -> Optional[float]:
            return self._value

    class MissingRatesDialog(QtWidgets.QDialog):
        """
        缺失币种：表格填空（外币/美元）。
        - 继续应用：返回 {CCY: rate}
        - 跳过：返回 {}
        """
        def __init__(self, missing_list: List[str], parent=None):
            super().__init__(parent)
            self.setWindowTitle("补充缺失汇率")
            self.setMinimumSize(560, 380)
            self.setStyleSheet(DIALOG_STYLE)
            self._rates: Dict[str, float] = {}
            self._skipped = False

            layout = QtWidgets.QVBoxLayout(self)
            layout.setContentsMargins(16, 16, 16, 16)
            layout.setSpacing(10)

            title = QtWidgets.QLabel("补充缺失汇率")
            f = QtGui.QFont()
            f.setPointSize(13)
            f.setWeight(font_bold_weight())
            title.setFont(f)
            layout.addWidget(title)

            hint = QtWidgets.QLabel(
                "rate 表未匹配到以下币种的“外币/美元”。\n"
                "请在表格中填写（例如 JPY 填 156.90）。留空表示跳过该币种。"
            )
            hint.setObjectName("hint")
            hint.setWordWrap(True)
            layout.addWidget(hint)

            self.table = QtWidgets.QTableWidget()
            self.table.setColumnCount(2)
            self.table.setHorizontalHeaderLabels(["CCY", "外币/美元"])
            self.table.setRowCount(len(missing_list))
            self.table.verticalHeader().setVisible(False)
            self.table.horizontalHeader().setStretchLastSection(True)

            for i, ccy in enumerate(missing_list):
                item_ccy = non_editable_item(QtWidgets.QTableWidgetItem(ccy))
                self.table.setItem(i, 0, item_ccy)
                self.table.setItem(i, 1, QtWidgets.QTableWidgetItem(""))

            self.table.resizeColumnsToContents()
            layout.addWidget(self.table, 1)

            btn_row = QtWidgets.QHBoxLayout()
            btn_row.addStretch(1)

            self.skip_btn = QtWidgets.QPushButton("跳过")
            self.skip_btn.setObjectName("secondary")
            self.ok_btn = QtWidgets.QPushButton("继续应用")
            self.ok_btn.setObjectName("primary")

            btn_row.addWidget(self.skip_btn)
            btn_row.addWidget(self.ok_btn)
            layout.addLayout(btn_row)

            self.skip_btn.clicked.connect(self._on_skip)
            self.ok_btn.clicked.connect(self._on_ok)

        def _on_skip(self):
            self._skipped = True
            self._rates = {}
            self.accept()

        def _on_ok(self):
            rates: Dict[str, float] = {}
            bad = []
            for r in range(self.table.rowCount()):
                ccy = (self.table.item(r, 0).text() or "").strip().upper()
                val_item = self.table.item(r, 1)
                txt = (val_item.text() if val_item else "").strip().replace(",", "").replace(" ", "")
                if not txt:
                    continue
                try:
                    v = float(txt)
                except ValueError:
                    bad.append(ccy)
                    continue
                if v <= 0:
                    bad.append(ccy)
                    continue
                rates[ccy] = float(v)

            if bad:
                QtWidgets.QMessageBox.warning(self, "提示", "以下币种输入不合法（需为正数）：\n" + ", ".join(bad))
                return

            self._rates = rates
            self.accept()

        def result_rates(self) -> Dict[str, float]:
            return dict(self._rates)

        def skipped(self) -> bool:
            return bool(self._skipped)

    class Worker(QtCore.QObject):
        progress = Signal(int)
        status = Signal(str)
        rates = Signal(object)

        request_threshold_confirm = Signal(object)  # payload dict
        request_missing_rates = Signal(object)      # list[str]

        finished = Signal(object)  # payload dict
        error = Signal(str)

        def __init__(self, file_path: str, manual_threshold: str, manual_rates_text: str):
            super().__init__()
            self.file_path = file_path
            self.manual_threshold = manual_threshold
            self.manual_rates_text = manual_rates_text

            self._threshold_event = threading.Event()
            self._threshold_value: Optional[float] = None
            self._threshold_cancelled = False

            self._missing_event = threading.Event()
            self._missing_rates: Dict[str, float] = {}

        @Slot(object)
        def set_threshold_result(self, payload: object):
            """
            payload: {"ok": bool, "value": float|None}
            """
            ok = False
            val = None
            if isinstance(payload, dict):
                ok = bool(payload.get("ok", False))
                val = payload.get("value", None)
            if not ok or val is None:
                self._threshold_cancelled = True
                self._threshold_value = None
            else:
                self._threshold_cancelled = False
                self._threshold_value = float(val)
            self._threshold_event.set()

        @Slot(object)
        def set_missing_rates_result(self, payload: object):
            """
            payload: {"rates": {CCY: rate}}
            """
            rates = {}
            if isinstance(payload, dict):
                rates = payload.get("rates", {}) or {}
            self._missing_rates = dict(rates)
            self._missing_event.set()

        def run(self):
            def status(msg: str):
                self.status.emit(msg)

            def prog(pct: int):
                self.progress.emit(int(max(0, min(100, pct))))

            try:
                prog(0)
                status("读取 rate 表...")

                # 1) 读取 rate（自动识别阈值 + 自动推导汇率）
                fx_rate, threshold_detected, rate_sheet, details = load_fx_rates_from_rate_sheet(
                    self.file_path, THRESHOLD_USD
                )

                # 合并主界面的手动汇率（覆盖/补充）
                manual_rates = parse_manual_rates(self.manual_rates_text or "")
                if manual_rates:
                    fx_rate.update(manual_rates)

                fx_rate.setdefault("USD", 1.0)

                # 手工阈值作为“当前值”，但仍然要弹窗确认（你要求“先识别再确认”）
                manual_threshold_val = parse_amount_input(self.manual_threshold or "")
                current_threshold = manual_threshold_val if manual_threshold_val else threshold_detected

                # 展示读取到的汇率（用于 UI 显示）
                self.rates.emit(
                    {
                        "rates": fx_rate,
                        "threshold_usd": threshold_detected,
                        "sheet_name": rate_sheet,
                        "source": details.get("source", "sheet"),
                    }
                )

                # 2) 阈值确认弹窗（可编辑）
                status("请确认阈值...")
                self._threshold_event.clear()
                self.request_threshold_confirm.emit(
                    {
                        "detected": threshold_detected,
                        "current": current_threshold,
                        "rate_sheet": rate_sheet,
                    }
                )
                self._threshold_event.wait()
                if self._threshold_cancelled or self._threshold_value is None:
                    raise ValueError("已取消：未确认阈值。")

                threshold_final = float(self._threshold_value)

                # 3) 预扫描缺失币种 -> 弹窗补录/跳过
                status("检查缺失币种汇率...")
                missing_list = prescan_missing_currencies(self.file_path, rate_sheet, fx_rate)
                skipped_ccy: Set[str] = set()

                if missing_list:
                    self._missing_event.clear()
                    self.request_missing_rates.emit(missing_list)
                    self._missing_event.wait()

                    provided = dict(self._missing_rates or {})
                    if provided:
                        fx_rate.update(provided)

                    skipped_ccy = set(missing_list) - set(provided.keys())

                # 4) 正式处理 & 导出
                status("开始筛选并导出...")
                prog(10)

                def progress_cb(p):
                    prog(p)

                def status_cb(msg):
                    status(msg)

                # 用 override 直接跑，避免重复读 rate
                count = process_workbook(
                    self.file_path,
                    progress_cb=progress_cb,
                    status_cb=status_cb,
                    rates_cb=None,
                    missing_cb=None,
                    fx_rate_override=fx_rate,
                    threshold_override=threshold_final,
                    rate_sheet_override=rate_sheet,
                    sheet_source_override=details.get("source", "sheet"),
                    suppress_missing_ccy=skipped_ccy,
                )

                prog(100)
                self.finished.emit(
                    {
                        "count": int(count),
                        "threshold": threshold_final,
                        "output_sheet": OUTPUT_SHEET_NAME,
                        "skipped_ccy": sorted(list(skipped_ccy)),
                    }
                )

            except Exception as exc:
                self.error.emit(str(exc))

    class MainWindow(QtWidgets.QWidget):
        threshold_result = Signal(object)      # -> worker.set_threshold_result
        missing_rates_result = Signal(object)  # -> worker.set_missing_rates_result

        def __init__(self):
            super().__init__()
            self._worker_thread = None
            self._worker = None
            self._build_ui()

        def _build_ui(self):
            self.setWindowTitle("USD 大额筛选")
            self.setMinimumSize(780, 540)
            self.resize(900, 600)

            if os.path.exists(ICON_PATH):
                self.setWindowIcon(QtGui.QIcon(ICON_PATH))

            base_font = QtGui.QFont("Avenir Next", 11)
            self.setFont(base_font)

            title_font = QtGui.QFont("Avenir Next", 22)
            title_font.setWeight(font_bold_weight())
            subtitle_font = QtGui.QFont("Avenir Next", 11)
            label_font = QtGui.QFont("Avenir Next", 11)
            button_font = QtGui.QFont("Avenir Next", 11)
            button_font.setWeight(font_bold_weight())
            small_font = QtGui.QFont("Avenir Next", 9)
            footer_font = QtGui.QFont("Avenir Next", 9)

            background = BackgroundWidget()
            root_layout = QtWidgets.QVBoxLayout(self)
            root_layout.setContentsMargins(0, 0, 0, 0)
            root_layout.addWidget(background)

            panel = QtWidgets.QFrame()
            panel.setObjectName("panel")
            panel.setAttribute(wa_styled, True)
            panel.setMaximumWidth(780)

            panel_layout = QtWidgets.QVBoxLayout(panel)
            panel_layout.setContentsMargins(26, 24, 26, 22)
            panel_layout.setSpacing(12)

            header_layout = QtWidgets.QHBoxLayout()
            header_layout.setSpacing(16)

            logo_pixmap = create_rounded_pixmap(LOGO_SRC_PATH, 86, 18)
            if logo_pixmap:
                logo_label = QtWidgets.QLabel()
                logo_label.setFixedSize(86, 86)
                logo_label.setPixmap(logo_pixmap)
                logo_label.setAlignment(align_center)
                header_layout.addWidget(logo_label, 0, align_vcenter)

            title_layout = QtWidgets.QVBoxLayout()
            title_label = QtWidgets.QLabel("USD 大额筛选")
            title_label.setFont(title_font)
            subtitle_label = QtWidgets.QLabel("自动识别 rate 表阈值与汇率；阈值确认后导出筛选结果")
            subtitle_label.setObjectName("subtitle")
            subtitle_label.setFont(subtitle_font)
            subtitle_label.setWordWrap(True)
            title_layout.addWidget(title_label)
            title_layout.addWidget(subtitle_label)
            header_layout.addLayout(title_layout, 1)

            panel_layout.addLayout(header_layout)

            file_label = QtWidgets.QLabel("Excel 文件")
            file_label.setFont(label_font)
            panel_layout.addWidget(file_label)

            file_row = QtWidgets.QHBoxLayout()
            file_row.setSpacing(12)
            self.path_edit = QtWidgets.QLineEdit()
            self.path_edit.setPlaceholderText("请选择 Excel 文件")
            file_row.addWidget(self.path_edit, 1)

            self.browse_btn = QtWidgets.QPushButton("选择文件")
            self.browse_btn.setObjectName("secondaryButton")
            self.browse_btn.setFont(button_font)
            self.browse_btn.clicked.connect(self.choose_file)
            file_row.addWidget(self.browse_btn)

            panel_layout.addLayout(file_row)

            hint_label = QtWidgets.QLabel(f"输出 Sheet：{OUTPUT_SHEET_NAME}（Transfer Amount 输出为绝对值）")
            hint_label.setObjectName("hint")
            hint_label.setFont(small_font)
            panel_layout.addWidget(hint_label)

            threshold_label = QtWidgets.QLabel("USD 阈值（可选：可留空，系统会先识别再弹窗确认）")
            threshold_label.setFont(label_font)
            panel_layout.addWidget(threshold_label)

            self.threshold_edit = QtWidgets.QLineEdit()
            self.threshold_edit.setPlaceholderText("留空则自动识别（例如：20M / 20000000）")
            panel_layout.addWidget(self.threshold_edit)

            manual_rate_label = QtWidgets.QLabel("手动汇率（可选：覆盖/补充，格式 JPY=156.90）")
            manual_rate_label.setFont(label_font)
            panel_layout.addWidget(manual_rate_label)

            self.manual_rate_box = QtWidgets.QPlainTextEdit()
            self.manual_rate_box.setObjectName("manualRateBox")
            self.manual_rate_box.setMinimumHeight(70)
            self.manual_rate_box.setMaximumHeight(110)
            self.manual_rate_box.setFont(small_font)
            panel_layout.addWidget(self.manual_rate_box)

            rate_label = QtWidgets.QLabel("已识别汇率（外币/美元）")
            rate_label.setFont(label_font)
            panel_layout.addWidget(rate_label)

            self.rate_box = QtWidgets.QPlainTextEdit()
            self.rate_box.setObjectName("rateBox")
            self.rate_box.setReadOnly(True)
            self.rate_box.setMinimumHeight(100)
            self.rate_box.setMaximumHeight(140)
            self.rate_box.setFont(small_font)
            self.rate_box.setPlainText("尚未读取。")
            panel_layout.addWidget(self.rate_box)

            action_row = QtWidgets.QHBoxLayout()
            self.run_btn = QtWidgets.QPushButton("开始筛选并导出")
            self.run_btn.setObjectName("primaryButton")
            self.run_btn.setFont(button_font)
            self.run_btn.clicked.connect(self.start_task)
            action_row.addWidget(self.run_btn, 0, align_left)
            action_row.addStretch(1)
            panel_layout.addLayout(action_row)

            self.progress_bar = QtWidgets.QProgressBar()
            self.progress_bar.setRange(0, 100)
            self.progress_bar.setValue(0)
            self.progress_bar.setTextVisible(True)
            panel_layout.addWidget(self.progress_bar)

            self.status_label = QtWidgets.QLabel("等待选择文件…")
            self.status_label.setObjectName("status")
            self.status_label.setFont(small_font)
            panel_layout.addWidget(self.status_label)

            footer_row = QtWidgets.QHBoxLayout()
            footer_row.addStretch(1)
            footer_label = QtWidgets.QLabel("i Designed by 余智秋 in Shanghai.")
            footer_label.setObjectName("footer")
            footer_label.setFont(footer_font)
            footer_row.addWidget(footer_label, 0, align_right)
            panel_layout.addLayout(footer_row)

            background_layout = QtWidgets.QVBoxLayout(background)
            background_layout.setContentsMargins(32, 28, 32, 28)
            background_layout.addStretch(1)
            background_layout.addWidget(panel, 0, align_center)
            background_layout.addStretch(1)

            panel.setStyleSheet(
                """
                QFrame#panel {
                    background-color: #0F212B;
                    border: 1px solid #173443;
                    border-radius: 18px;
                }
                QLabel { color: #E6F1F5; }
                QLabel#subtitle, QLabel#hint, QLabel#status { color: #9FB2BC; }
                QLabel#footer { color: #D6B25E; }
                QLineEdit {
                    background: #0A1720;
                    border: 1px solid #23414D;
                    border-radius: 10px;
                    padding: 8px 10px;
                    color: #E6F1F5;
                }
                QLineEdit:focus { border-color: #44E6C2; }
                QPlainTextEdit#rateBox, QPlainTextEdit#manualRateBox {
                    background: #0A1720;
                    border: 1px solid #23414D;
                    border-radius: 10px;
                    padding: 6px 8px;
                    color: #E6F1F5;
                }
                QPushButton#primaryButton {
                    background-color: #44E6C2;
                    color: #0A1B22;
                    border-radius: 10px;
                    padding: 8px 18px;
                }
                QPushButton#primaryButton:hover { background-color: #3BD9B7; }
                QPushButton#primaryButton:pressed { background-color: #2EB596; }
                QPushButton#secondaryButton {
                    background-color: #132833;
                    color: #E6F1F5;
                    border: 1px solid #23414D;
                    border-radius: 10px;
                    padding: 8px 16px;
                }
                QPushButton#secondaryButton:hover { border-color: #44E6C2; }
                QProgressBar {
                    background: #0A1720;
                    border: 1px solid #23414D;
                    border-radius: 6px;
                    text-align: center;
                    color: #9FB2BC;
                }
                QProgressBar::chunk {
                    background: #44E6C2;
                    border-radius: 6px;
                }
                """
            )

        def choose_file(self):
            file_path, _ = QtWidgets.QFileDialog.getOpenFileName(
                self, "选择 Excel 文件", BASE_DIR, "Excel 文件 (*.xlsx *.xlsm *.xls)"
            )
            if file_path:
                self.path_edit.setText(file_path)
                self.status_label.setText("已选择文件，准备开始。")

        def _set_running(self, running: bool):
            self.run_btn.setEnabled(not running)
            self.browse_btn.setEnabled(not running)
            self.path_edit.setEnabled(not running)

        def start_task(self):
            file_path = self.path_edit.text().strip()
            if not file_path:
                QtWidgets.QMessageBox.warning(self, "提示", "请先选择 Excel 文件。")
                return
            if not os.path.exists(file_path):
                QtWidgets.QMessageBox.critical(self, "错误", "文件不存在，请重新选择。")
                return

            manual_threshold_text = self.threshold_edit.text().strip()
            if manual_threshold_text and parse_amount_input(manual_threshold_text) is None and _to_num(manual_threshold_text) is None:
                QtWidgets.QMessageBox.warning(self, "提示", "USD 阈值格式不正确，请重新输入。")
                return

            manual_rates_text = self.manual_rate_box.toPlainText().strip()
            if manual_rates_text:
                parsed_manual = parse_manual_rates(manual_rates_text)
                if not parsed_manual:
                    QtWidgets.QMessageBox.warning(self, "提示", "手动汇率格式不正确，请按示例填写：JPY=156.90")
                    return

            self.progress_bar.setValue(0)
            self.status_label.setText("准备处理中…")
            self.rate_box.setPlainText("读取 rate 中…")
            self._set_running(True)

            self._worker_thread = QtCore.QThread()
            self._worker = Worker(file_path, manual_threshold_text, manual_rates_text)
            self._worker.moveToThread(self._worker_thread)

            # UI -> Worker 回传
            self.threshold_result.connect(self._worker.set_threshold_result)
            self.missing_rates_result.connect(self._worker.set_missing_rates_result)

            # Worker -> UI
            self._worker_thread.started.connect(self._worker.run)
            self._worker.progress.connect(self.progress_bar.setValue)
            self._worker.status.connect(self.status_label.setText)
            self._worker.rates.connect(self._handle_rates)
            self._worker.request_threshold_confirm.connect(self._confirm_threshold_dialog)
            self._worker.request_missing_rates.connect(self._missing_rates_dialog)
            self._worker.finished.connect(self._handle_finished)
            self._worker.error.connect(self._handle_error)

            self._worker.finished.connect(self._worker_thread.quit)
            self._worker.error.connect(self._worker_thread.quit)
            self._worker_thread.finished.connect(self._cleanup_worker)

            self._worker_thread.start()

        def _confirm_threshold_dialog(self, payload: dict):
            try:
                detected = float(payload.get("detected", THRESHOLD_USD))
                current = float(payload.get("current", detected))
                rate_sheet = str(payload.get("rate_sheet", RATE_SHEET_NAME))
            except Exception:
                detected, current, rate_sheet = float(THRESHOLD_USD), float(THRESHOLD_USD), RATE_SHEET_NAME

            dlg = ThresholdConfirmDialog(detected=detected, current=current, rate_sheet=rate_sheet, parent=self)
            if hasattr(dlg, "exec"):
                ok = dlg.exec()
            else:
                ok = dlg.exec_()

            if ok:
                val = dlg.value()
                self.threshold_edit.setText(f"{val:,.2f}" if val is not None else "")
                self.threshold_result.emit({"ok": True, "value": val})
            else:
                self.threshold_result.emit({"ok": False, "value": None})

        def _missing_rates_dialog(self, missing_list: list):
            missing = [str(x).strip().upper() for x in (missing_list or []) if str(x).strip()]
            dlg = MissingRatesDialog(missing, parent=self)
            if hasattr(dlg, "exec"):
                dlg.exec()
            else:
                dlg.exec_()
            rates = dlg.result_rates()
            self.missing_rates_result.emit({"rates": rates})

            # 同步到主界面的手动汇率输入框，方便用户后续复用/留档
            if rates:
                existing = self.manual_rate_box.toPlainText().strip()
                lines = []
                if existing:
                    lines.append(existing)
                for ccy, v in sorted(rates.items()):
                    lines.append(f"{ccy}={v}")
                self.manual_rate_box.setPlainText("\n".join(lines))

        def _handle_rates(self, rates: dict):
            if not rates or not isinstance(rates, dict):
                self.rate_box.setPlainText("未读取到汇率。")
                return
            rate_map = rates.get("rates") or {}
            threshold_detected = float(rates.get("threshold_usd", THRESHOLD_USD))
            sheet_name = rates.get("sheet_name", RATE_SHEET_NAME)
            source = rates.get("source", "sheet")

            source_text = "来源：EXCHANGE VS CNY 推导" if source == "exchange_vs_cny" else "来源：LIMIT AMOUNT 推导"

            lines = [
                f"自动识别阈值（待确认）：{threshold_detected:,.2f}",
                f"rate 表：{sheet_name}",
                source_text,
                "汇率（外币/美元）：",
            ]
            for ccy, rate in sorted(rate_map.items()):
                try:
                    lines.append(f"{ccy}: {float(rate):.6f}")
                except Exception:
                    lines.append(f"{ccy}: {rate}")
            self.rate_box.setPlainText("\n".join(lines))

        def _handle_finished(self, payload: dict):
            count = int(payload.get("count", 0))
            threshold = float(payload.get("threshold", THRESHOLD_USD))
            out_sheet = payload.get("output_sheet", OUTPUT_SHEET_NAME)
            skipped = payload.get("skipped_ccy", []) or []

            msg = f"已完成导出：{out_sheet}\n阈值：{threshold:,.2f}\n命中记录：{count}"
            if skipped:
                msg += "\n\n已跳过缺失汇率币种：\n" + ", ".join(skipped)

            QtWidgets.QMessageBox.information(self, "完成", msg)
            self.status_label.setText("处理完成。")
            self._set_running(False)
            open_file_path(self.path_edit.text().strip())

        def _handle_error(self, message: str):
            QtWidgets.QMessageBox.critical(self, "错误", message)
            self.status_label.setText("处理失败。")
            self._set_running(False)

        def _cleanup_worker(self):
            if self._worker is not None:
                self._worker.deleteLater()
                self._worker = None
            if self._worker_thread is not None:
                self._worker_thread.deleteLater()
                self._worker_thread = None

    app = QtWidgets.QApplication(sys.argv)
    try:
        app.setStyle("Fusion")
    except Exception:
        pass

    window = MainWindow()
    window.show()

    if hasattr(app, "exec"):
        app.exec()
    else:
        app.exec_()
    return True


def main():
    if not launch_ui():
        raise SystemExit(1)


if __name__ == "__main__":
    main()
